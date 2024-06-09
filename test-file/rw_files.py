# 读写文件测试，包括log/txt/json/xml/csv/excel/word等文件类型
import csv
import json
import os
import random
import xml.etree.cElementTree as ET

import win32com
import win32com.client
from openpyxl import Workbook, load_workbook


def check_path(file_path):
    """检查目标文件是否存在，不存在新建"""
    if not os.path.exists(file_path):
        with open(file_path, 'x', encoding='utf-8'):
            print(f'file path not exists, create new file ----> {file_path}')


def test_operate_log(file_path, file_type):
    check_path(file_path)
    with open(file_path, 'a+', encoding='utf-8') as fw:
        random_num = random.randint(1, 1000)
        fw.writelines(str(random_num))
        print(f'write {random_num} to {file_path}')
    with open(file_path, 'r') as fr:
        print(f'output {file_type} ....... {fr.readlines()}')


def test_operate_json2(file_path, file_type):
    check_path(file_path)
    dicts = {'a': True, 'b': 3.1415, 'c': "testing"}
    # 写入json文件
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(dicts, f)
        print(f'write {dicts} to {file_path}')
    # 读取json文件
    with open(file_path, 'r') as fr:
        print(f'output {file_type} ....... {json.load(fr)}')


def test_operate_csv(file_path, file_type):
    check_path(file_path)
    # 按行写入csv文件, 每个列表是一行
    content = [[1, 2, 3, 4, 5, 6, 7, 8, 9],
               ["china", "china", "china", "china", "china", "china", "china", "china"],
               [3, 3.14, True, 'China']]
    # 写入操作：newline是指定新一行分隔符，不写newline会隔一行写一行
    with open(file_path, 'w', encoding='utf-8', newline="") as fw:
        csv.writer(fw).writerows(content)
    # 读取操作：
    with open(file_path, 'r', encoding='utf-8') as fr:
        print(f'output {file_type} .............')
        for line in csv.reader(fr):
            print(f'    ....... : {line}')


def test_operate_xml(file_path, file_type):
    check_path(file_path)
    # 写入操作：
    with open(file_path, 'w', encoding='utf-8') as fw:
        lines = '<project>\n' \
                '\t<modelVersion>4.0.0</modelVersion>\n' \
                '\t<packaging>pom</packaging>\n' \
                '</project>'
        fw.writelines(lines)
    # 读取操作：
    root = ET.parse(file_path).getroot()
    print(f'output {file_type} root ....... {root}')
    for parent_child in root:
        print(f'output {file_type} parent_child ....... parent-tag：{parent_child.tag}, parent-text：{parent_child.text}')


def test_operate_excel(file_path, file_type):
    check_path(file_path)
    # 写入Excel操作：
    print(f"开始写入Excel ....... {file_path}")
    wb = Workbook()  # 打开工作簿
    sheet = wb.active  # 激活sheet
    t_headers = ['学号', '姓名', '专业', '班级', '住宿地址']  # 表头
    for i in range(len(t_headers)):
        sheet.cell(row=1, column=i + 1).value = t_headers[i]
    t_bodys = [
        ['1001', '王同学', '电子信息工程', '1班', '28-429'],
        ['1002', '周同学', '历史系', '1班', '1-101'],
        ['1003', '张同学', '艺术系', '1班', '5-205']
    ]
    for j in range(len(t_bodys)):
        sheet.append(t_bodys[j])
    # 保存Excel文件
    wb.save(file_path)
    wb.close()
    print(f"结束写入Excel ....... {file_path}")
    # 读取Excel操作：
    print(f"开始读取Excel ....... {file_path}")
    lw = load_workbook(file_path)
    sheet_names = lw.sheetnames
    print(f"Excel sheet_names ....... {str(sheet_names)}")
    # 遍历sheet
    for sheet_name in sheet_names:
        current_sheet = lw[sheet_name]
        print(f"Excel current_sheet ....... "
              f"title: {current_sheet.title}, max_row: {current_sheet.max_row}, max_column: {current_sheet.max_column}")
        # 按行读取,不包含表头
        line_list = []
        for line in range(2, current_sheet.max_row + 1):
            for col in range(1, current_sheet.max_column + 1):
                line_list.append(current_sheet.cell(row=line, column=col).value)
        print(f"Excel sheet_name ....... {str(sheet_names)}, result: {str(line_list)}")
    print(f"结束读取Excel ....... {file_path}")


def test_operate_word(file_path, file_type):
    check_path(file_path)
    # 写入word操作：
    print(f"开始写入word ....... {file_path}")
    w_app = win32com.client.Dispatch('Word.Application')  # 打开office
    w_app.Visible = True  # 设置可见
    # 创建文档
    doc = w_app.Documents.Add()
    start = doc.Range(0, 0)  # 文本开始位置
    start.InsertBefore('Hello Word\n')
    start.InsertAfter('我是weixiangxiang\n')
    start.InsertAfter('我在测试数据写入Word文档操作\n')
    end = doc.Range()  # 文本结束位置
    end.InsertAfter('updatetime: 20240606')
    # 第二个参数，不写默认docx，传2则为txt（转储到txt）
    doc.SaveAs(file_path)
    # 关闭，True-保存内容，False则不保存
    doc.Close(True)
    w_app.Quit()
    print(f"结束写入word ....... {file_path}")
    # 读取word操作：
    print(f"开始读取word ....... {file_path}")
    r_app = win32com.client.Dispatch('Word.Application')
    word = r_app.Documents.Open(file_path)
    for w in word.Paragraphs:
        print(f'output {file_type} line text ....... {w.Range.Text}')
    # 关闭
    word.Close()
    r_app.Quit()
    print(f"结束读取word ....... {file_path}")


if __name__ == '__main__':
    base_path = 'D:\\My Test\\XXX\\test-file\\'
    type_list = ['log', 'txt', 'json', 'csv', 'xml', 'xlsx', 'docx']
    test_operate_log(base_path + "test1.log", type_list[0])
    test_operate_log(base_path + "test2.txt", type_list[1])
    test_operate_json2(base_path + "test3.json", type_list[2])
    test_operate_csv(base_path + "test4.csv", type_list[3])
    test_operate_xml(base_path + "test5.xml", type_list[4])
    test_operate_excel(base_path + "test6.xlsx", type_list[5])
    test_operate_word(base_path + "test7.docx", type_list[6])
